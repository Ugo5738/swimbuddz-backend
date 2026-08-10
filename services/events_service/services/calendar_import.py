"""Parse the controlled Calendar Import worksheet into canonical event drafts."""

import re
from datetime import date, datetime, time
from io import BytesIO
from typing import Any, Optional
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from services.events_service.schemas.planning import (
    CalendarImportEvent,
    CalendarImportPreviewItem,
    CalendarImportPreviewResponse,
)

SHEET_NAME = "Calendar Import"
MAX_IMPORT_ROWS = 500
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
REQUIRED_HEADERS = {
    "import",
    "start date",
    "start time",
    "end date",
    "end time",
    "title",
    "audience",
    "visibility",
    "tier access",
    "event type",
    "location type",
    "timezone",
    "external key",
}


def _key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _key(value)).strip("_")


def _as_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Use a valid Excel date")


def _as_time(value: Any) -> time:
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if isinstance(value, (float, int)) and 0 <= float(value) < 1:
        minutes = round(float(value) * 24 * 60)
        return time(hour=(minutes // 60) % 24, minute=minutes % 60)
    text = str(value or "").strip()
    for pattern in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(text, pattern).time()
        except ValueError:
            continue
    raise ValueError("Use a valid Excel time")


def _map_audience(value: Any) -> tuple[str, Optional[str]]:
    text = _key(value)
    if "academy" in text and "club" not in text:
        return "academy", None
    if "club" in text and "community" not in text and "academy" not in text:
        return "club", None
    if "academy" in text and "club" in text:
        return (
            "academy",
            "Combined Academy/Club audience is shown in the Academy calendar lane.",
        )
    if "community" in text and "club" in text:
        return (
            "community",
            "Community includes Club members; the event uses the Community lane.",
        )
    if text in {"all", "everyone", "full community", "community"}:
        return "community", None
    raise ValueError("Audience must identify Community, Club, or Academy")


def _map_visibility(value: Any) -> str:
    text = _slug(value)
    mapping = {
        "public": "public",
        "members_only": "members_only",
        "member_only": "members_only",
        "invite_only": "invite_only",
    }
    if text not in mapping:
        raise ValueError("Visibility must be Public, Members-only, or Invite-only")
    return mapping[text]


def _map_tier_access(value: Any, visibility: str) -> tuple[str, Optional[str]]:
    if visibility == "invite_only":
        return "invite_only", None
    text = _key(value)
    if text in {"all", "public", "everyone", "anyone"}:
        return "public", None
    if "community" in text:
        return "community", None
    if "club" in text and "academy" in text:
        return (
            "community",
            "Club + Academy access is represented as Community access until multi-tier OR access is supported.",
        )
    if text == "club":
        return "club", None
    if text == "academy":
        return "academy", None
    raise ValueError(
        "Tier Access must be All, Community, Club, Academy, or Invite-only"
    )


def _map_location_type(value: Any) -> str:
    text = _slug(value)
    mapping = {"online": "online", "physical": "physical", "hybrid": "hybrid"}
    if text not in mapping:
        raise ValueError("Location Type must be Online, Physical, or Hybrid")
    return mapping[text]


def _map_pricing_mode(value: Any, attendee_price: Any) -> str:
    text = _slug(value)
    if not text:
        return "fixed" if attendee_price not in {None, ""} else "free"
    mapping = {
        "free": "free",
        "included": "included",
        "included_in_membership": "included",
        "fixed": "fixed",
        "fixed_price": "fixed",
        "cost_plus": "cost_plus",
        "calculated": "cost_plus",
    }
    if text not in mapping:
        raise ValueError("Pricing Mode must be Free, Included, Fixed, or Cost plus")
    return mapping[text]


def _map_reminder_hours(value: Any, event_type: str) -> list[int]:
    text = _slug(value)
    if not text:
        return [168, 24, 1] if event_type == "online_talk" else []
    profiles = {
        "none": [],
        "standard": [72, 24],
        "online_talk": [168, 24, 1],
        "major_event": [336, 168, 24],
    }
    if text not in profiles:
        raise ValueError(
            "Reminder Profile must be None, Standard, Online talk, or Major event"
        )
    return profiles[text]


def parse_calendar_import(content: bytes) -> CalendarImportPreviewResponse:
    try:
        with ZipFile(BytesIO(content)) as archive:
            if (
                sum(entry.file_size for entry in archive.infolist())
                > MAX_UNCOMPRESSED_BYTES
            ):
                raise ValueError("Workbook expands beyond the 50 MB safety limit")
    except BadZipFile as exc:
        raise ValueError("The uploaded file is not a valid .xlsx workbook") from exc

    # Some spreadsheet generators omit the optional worksheet dimension hint.
    # Normal mode derives dimensions from the cells and is bounded by the 5 MB
    # upload limit enforced by the route.
    try:
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError) as exc:
        raise ValueError("The uploaded file is not a valid .xlsx workbook") from exc
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f'Workbook must include a "{SHEET_NAME}" worksheet')
    sheet = workbook[SHEET_NAME]

    header_row = None
    header_map: dict[str, int] = {}
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        candidate = {
            _key(sheet.cell(row_index, column_index).value): column_index
            for column_index in range(1, sheet.max_column + 1)
            if sheet.cell(row_index, column_index).value is not None
        }
        if REQUIRED_HEADERS.issubset(candidate):
            header_row = row_index
            header_map = candidate
            break
    if header_row is None:
        missing = ", ".join(sorted(REQUIRED_HEADERS))
        raise ValueError(f"Calendar Import headers were not found. Required: {missing}")

    def value(row: int, name: str) -> Any:
        column = header_map.get(name)
        return sheet.cell(row, column).value if column else None

    items: list[CalendarImportPreviewItem] = []
    for row in range(
        header_row + 1, min(sheet.max_row, header_row + MAX_IMPORT_ROWS) + 1
    ):
        include = _key(value(row, "import")) in {"yes", "y", "true", "1"}
        if not include and not any(
            value(row, name) for name in REQUIRED_HEADERS - {"import"}
        ):
            continue
        errors: list[str] = []
        warnings: list[str] = []
        event = None
        try:
            timezone_name = str(value(row, "timezone") or "Africa/Lagos").strip()
            try:
                timezone = ZoneInfo(timezone_name)
            except ZoneInfoNotFoundError as exc:
                raise ValueError("Timezone is not recognized") from exc
            start_date = _as_date(value(row, "start date"))
            end_date = _as_date(value(row, "end date") or start_date)
            start_time = _as_time(value(row, "start time"))
            end_time = _as_time(value(row, "end time"))
            starts_at = datetime.combine(start_date, start_time, tzinfo=timezone)
            ends_at = datetime.combine(end_date, end_time, tzinfo=timezone)
            if ends_at <= starts_at:
                raise ValueError("End date/time must be after start date/time")

            audience, audience_warning = _map_audience(value(row, "audience"))
            if audience_warning:
                warnings.append(audience_warning)
            visibility = _map_visibility(value(row, "visibility"))
            tier_access, access_warning = _map_tier_access(
                value(row, "tier access"), visibility
            )
            if access_warning:
                warnings.append(access_warning)

            location = str(value(row, "location") or "").strip() or None
            if location and "tbc" in location.lower():
                warnings.append("Confirm the venue or meeting link before publishing.")
            title = str(value(row, "title") or "").strip()
            if not title:
                raise ValueError("Title is required")
            external_key = str(value(row, "external key") or "").strip()
            if not external_key:
                raise ValueError("External Key is required for duplicate protection")

            event_type = _slug(value(row, "event type"))
            attendee_price = value(row, "attendee price")
            if attendee_price in {None, ""}:
                cost_naira = None
            else:
                cost_naira = float(attendee_price)
                if cost_naira < 0:
                    raise ValueError("Attendee Price cannot be negative")

            event = CalendarImportEvent(
                title=title,
                description=str(value(row, "description") or "").strip() or None,
                event_type=event_type,
                audience=audience,
                visibility=visibility,
                status="draft",
                location_type=_map_location_type(value(row, "location type")),
                timezone=timezone_name,
                location_area=str(value(row, "location area") or "").strip() or None,
                is_location_private=visibility == "invite_only",
                location=location,
                start_time=starts_at,
                end_time=ends_at,
                max_capacity=None,
                tier_access=tier_access,
                pool_id=None,
                cost_naira=cost_naira,
                pricing_mode=_map_pricing_mode(
                    value(row, "pricing mode"), attendee_price
                ),
                email_reminder_hours=_map_reminder_hours(
                    value(row, "reminder profile"), event_type
                ),
                external_key=external_key,
                source_sheet=str(value(row, "source sheet") or SHEET_NAME),
                source_row=int(value(row, "source row") or row),
            )
        except (TypeError, ValueError) as exc:
            errors.append(str(exc))
        items.append(
            CalendarImportPreviewItem(
                source_row=row,
                selected=include,
                event=event,
                warnings=warnings,
                errors=errors,
            )
        )

    return CalendarImportPreviewResponse(
        sheet_name=SHEET_NAME,
        valid_count=sum(
            1 for item in items if item.event is not None and not item.errors
        ),
        invalid_count=sum(1 for item in items if item.errors),
        rows=items,
    )
